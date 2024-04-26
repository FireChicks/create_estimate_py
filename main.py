import sys
import re
import os

from PyQt5.QtGui import QDragEnterEvent, QDropEvent
from PyQt5.QtWidgets import *
from PyQt5 import uic
from openpyxl import load_workbook
from collections import deque

import pandas as pd
import time

# ui 파일 불러오기
form_class = uic.loadUiType("create_estimate.ui")[0]
# 파일 가져오기 다이얼로그 생성

dict_dtype = {
    "행" : str,
    "문단" : str
}

add_cols = ['시트', '열', '행', '문단', '시트.1', '열.1', '행.1']

class WindowClass(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.selectedItems = []
        self.deleteRows = set()
        self.table_states = deque()
        self.current_state = {}
        self.first_state = {}

        self.suffix = '입력완료'
        self.suffixNum = ['_(%d)', '_%d', '%d']
        self.sufNum = 0
        self.num = 0

        self.transAddFilePath = ""
        self.transInpFilePath = ""
        self.transOutFilePath = ""

        self.addEditAddFilePath = ""

        self.transAddDF = pd.DataFrame()
        self.transInpDF = pd.DataFrame()
        self.transOutDF = pd.DataFrame()

        self.file_dialog = QFileDialog()

        #세팅 페이지 저장 문구 초기값
        self.set_suffix_lbl()

        # btn_move_exe 버튼에 대한 시그널-슬롯 연결
        self.btn_move_exe.clicked.connect(self.move_to_exe_page)
        self.btn_move_trans.clicked.connect(self.move_to_create_page)
        self.btn_move_setting.clicked.connect(self.move_to_setting_page)

        #메인 페이지 돌아가는 버튼 함수 연결
        self.btn_return_main.clicked.connect(self.return_main_page)
        self.btn_return_main_2.clicked.connect(self.return_main_page)
        self.btn_return_main_3.clicked.connect(self.return_main_page)

        #변환 페이지 버튼
        self.btn_set_trans.clicked.connect(self.load_file_add)
        self.btn_set_input.clicked.connect(self.load_file_inp)
        self.btn_set_output.clicked.connect(self.load_file_out)

        #drag & drop 관련 함수
        self.input_file_name.setAcceptDrops(True)
        self.input_file_name.dragEnterEvent = self.drag_enter_event
        self.input_file_name.dropEvent = self.drop_inp_event

        self.output_file_name.setAcceptDrops(True)
        self.output_file_name.dragEnterEvent = self.drag_enter_event
        self.output_file_name.dropEvent = self.drop_out_event

        self.trans_file_name.setAcceptDrops(True)
        self.trans_file_name.dragEnterEvent = self.drag_enter_event
        self.trans_file_name.dropEvent = self.drop_trans_event


        self.btn_exe.clicked.connect(self.trans_exe)


        #create 페이지 버튼들
        self.lbl_selected_file.setAcceptDrops(True)
        self.lbl_selected_file.dragEnterEvent = self.drag_enter_event
        self.lbl_selected_file.dropEvent = self.drop_load_file_event

        self.btn_add_row.clicked.connect(self.add_row)
       # self.btn_copy_row.clicked.connect(self.copy_row)
        self.btn_del_row.clicked.connect(self.del_row)
        self.tbl_trans.cellClicked.connect(self.add_clicked_row)
        self.tbl_trans.itemChanged.connect(self.save_table_state)

        self.btn_output_trans.clicked.connect(self.save_to_excel)
        self.btn_return_previous.clicked.connect(self.undo)
        
        #address파일 가져오기
        self.btn_input_trans.clicked.connect(self.load_file_dialog)


        #설정 페이지
        self.btn_chn_suffix.clicked.connect(self.change_suffix)

        self.btn_chn_suf_num.clicked.connect(self.change_sufNum)
        self.btn_chn_suf_num_2.clicked.connect(self.change_sufNum_2)
        self.btn_chn_suf_num_3.clicked.connect(self.change_sufNum_3)

    def move_to_exe_page(self):
        # QStackedWidget의 currentIndex를 변경하여 페이지를 전환
        self.stackedWidget.setCurrentIndex(1)

    def move_to_create_page(self):
        self.stackedWidget.setCurrentIndex(2)

    def move_to_setting_page(self):
        self.stackedWidget.setCurrentIndex(3)

    def return_main_page(self):
        self.stackedWidget.setCurrentIndex(0)

    def load_file_add(self):
        # 파일 다이얼로그 열기
        self.transAddFilePath, _ = self.file_dialog.getOpenFileName(self, "주소 지정 파일 가져오기", "",
                                                                      "All Files (*);;CSV Files (*.csv);;Excel Files (*.xlsx)")

        # 파일 경로를 QLabel에 표시
        if self.transAddFilePath:
            self.trans_file_name.setText(self.transAddFilePath.split('/')[-1])
        else:
            return

        # 파일을 데이터프레임으로 읽어오기
        df = pd.read_excel(self.transAddFilePath , skiprows=1, dtype=dict_dtype)
        df_drop_first_column = df.drop(df.columns[0], axis=1)
        self.transAddDF = df_drop_first_column.dropna(axis=0)

    def load_file_inp(self):
        # 파일 다이얼로그 열기
        self.transInpFilePath , _ = self.file_dialog.getOpenFileName(self, "입력 파일 가져오기", "",
                                                                      "All Files (*);;CSV Files (*.csv);;Excel Files (*.xlsx)")

        # 파일 경로를 QLabel에 표시
        if self.transInpFilePath:
            self.input_file_name.setText(self.transInpFilePath.split('/')[-1])

    def load_file_out(self):
        # 파일 다이얼로그 열기
        self.transOutFilePath , _ = self.file_dialog.getOpenFileName(self, "출력 파일 가져오기", "",
                                                                      "All Files (*);;CSV Files (*.csv);;Excel Files (*.xlsx)")

        # 파일 경로를 QLabel에 표시
        if self.transOutFilePath :
            self.output_file_name.setText(self.transOutFilePath.split('/')[-1])

        #예상 출력명 계산
        self.ex_output_name()

    #드래그 앤 드롭 관련 함수
    def drag_enter_event(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def drop_trans_event(self, event: QDropEvent):
        urls = [url.toLocalFile() for url in event.mimeData().urls()]
        #개수 검사
        if len(urls) > 1 :
            QMessageBox.warning(self, '너무 많은 개수의 파일 입력', '한가지 파일만 입력해주세요', QMessageBox.Ok)
            return

        if urls:
            url = urls[0]
            print("Dropped files:", url)
            #xlsx 파일인지 확인
            if url.split('.')[-1] != 'xlsx' :
                QMessageBox.warning(self, '지정되지 않은 파일 입력', 'xlsx 파일만 입력해주세요.', QMessageBox.Ok)
                return
            else:
                self.transAddFilePath = url
                self.trans_file_name.setText(self.transAddFilePath.split('/')[-1])

                # 파일을 데이터프레임으로 읽어오기
                df = pd.read_excel(self.transAddFilePath, skiprows=1, dtype=dict_dtype)
                df_drop_first_column = df.drop(df.columns[0], axis=1)
                self.transAddDF = df_drop_first_column.dropna(axis=0)

    def drop_inp_event(self, event: QDropEvent):
        urls = [url.toLocalFile() for url in event.mimeData().urls()]
        #개수 검사
        if len(urls) > 1 :
            QMessageBox.warning(self, '너무 많은 개수의 파일 입력', '한가지 파일만 입력해주세요', QMessageBox.Ok)
            return

        if urls:
            url = urls[0]
            print("Dropped files:", url)
            #xlsx 파일인지 확인
            if url.split('.')[-1] != 'xlsx' :
                QMessageBox.warning(self, '지정되지 않은 파일 입력', 'xlsx 파일만 입력해주세요.', QMessageBox.Ok)
                return
            else:
                self.transInpFilePath = url
                self.input_file_name.setText(self.transInpFilePath.split('/')[-1])

    def drop_out_event(self, event: QDropEvent):
        urls = [url.toLocalFile() for url in event.mimeData().urls()]
        #개수 검사
        if len(urls) > 1 :
            QMessageBox.warning(self, '너무 많은 개수의 파일 입력', '한가지 파일만 입력해주세요', QMessageBox.Ok)
            return

        if urls:
            url = urls[0]
            print("Dropped files:", url)
            #xlsx 파일인지 확인
            if url.split('.')[-1] != 'xlsx' :
                QMessageBox.warning(self, '지정되지 않은 파일 입력', 'xlsx 파일만 입력해주세요.', QMessageBox.Ok)
                return
            else:
                self.transOutFilePath = url
                self.output_file_name.setText(self.transOutFilePath.split('/')[-1])
                # 예상 출력명 계산
                self.ex_output_name()


        # 파일 놓기
    def dropEvent(self, event):
        urls = self.find_pdf(event.mimeData())
        if urls:
            for url in urls:
                print(url.toLocalFile())
            event.accept()
        else:
            event.ignore()

    def trans_exe(self):
        if self.transAddFilePath == '':
            QMessageBox.warning(self, '선택된 지정 파일이 없습니다.', '버튼을 눌러 파일을 지정해주세요.', QMessageBox.Ok)
            return
        elif self.transInpFilePath == '':
            QMessageBox.warning(self, '선택된 입력 파일이 없습니다.', '버튼을 눌러 파일을 지정해주세요.', QMessageBox.Ok)
            return
        elif self.transOutFilePath == '':
            QMessageBox.warning(self, '선택된 입력 파일이 없습니다.', '버튼을 눌러 파일을 지정해주세요.', QMessageBox.Ok)
            return
        else:
            result = []
            try:
                outputFileName = self.transOutFilePath.split('/')[-1].split('.')[0] + self.create_suffix()
                outputFileDir = ''

                #경로 구하는 과정
                for dir in self.transOutFilePath.split('/')[0:-1]:
                    outputFileDir  = outputFileDir + dir + '/'
                outputFileDir = outputFileDir + outputFileName

                opWb = load_workbook(self.transOutFilePath)
                opWb.save(outputFileDir)
                for idx, row in self.transAddDF.iterrows():
                    ipWb = load_workbook(self.transInpFilePath)
                    opWb = load_workbook(outputFileDir)

                    ipWs = ipWb[row['시트']]
                    try:
                        opWs = opWb[row['시트.1']]
                    except KeyError:
                        print(row['시트.1'] + " 시트가 존재하지 않습니다.")
                        continue

                    ipIdx = row['열'] + row['행']
                    opIdx = row['열.1'] + row['행.1']
                    nIdx = row['문단']

                    inputValue = ''
                    inputValues = str(ipWs[ipIdx].value).split('\n')

                    if len(inputValues) > 1:
                        inputValue = inputValues[int(nIdx) - 1]
                    else:
                        inputValue = inputValues[0]

                    # INPUT에 값이 없을때는 처리하지 않음
                    if inputValue != None :
                        opWs[opIdx] = str(inputValue)

                    opWb.save(outputFileDir)

                opWb.save(outputFileDir)
                self.lbl_result_txt.setText("성공적으로 파일 " + outputFileName.split('/')[-1] + "를 저장하였습니다.")
                self.ex_output_name()

            except Exception as e:
                print(e)
                QMessageBox.warning(self, '경고', '예상하지 못한 오류가 발생하였습니다.', QMessageBox.Ok)
                return


    #create 페이지 관련 함수들

    def save_table_state(self):
        #테이블의 상태를 저장하는 함수
        # 현재 테이블 상태를 저장
        current_state = {
            'data': [],
            'row_count': self.tbl_trans.rowCount(),
            'column_count': self.tbl_trans.columnCount()
        }
        for row in range(current_state['row_count']):
            row_data = []
            for col in range(current_state['column_count']):
                item = self.tbl_trans.item(row, col)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append('')
            current_state['data'].append(row_data)

        if len(self.table_states) > 0:
            if  (self.table_states[-1] != current_state) :
                if self.current_state != current_state:
                    self.table_states.append(self.current_state)
                    self.current_state = current_state
            else:
                return
        else:
            self.table_states.append(current_state)
            self.first_state = current_state


    def undo(self):
        #이전 상태로 돌아가는 함수
        if self.table_states:
            # 스택에서 최근 저장된 상태를 꺼내서 복원
            self.current_state = self.table_states.pop()

            if not self.first_state:
                QMessageBox.warning(self, '더이상 뒤돌아갈 행이 없습니다.', '행을 추가해주세요.', QMessageBox.Ok)

            if self.current_state != self.first_state:
                self.restore_table_state(self.current_state)
            else:
                self.restore_table_state(self.current_state)
                self.table_states.clear()
                self.first_state.clear()

            return
        else:
            QMessageBox.warning(self, '더이상 뒤돌아갈 행이 없습니다.', '행을 추가해주세요.', QMessageBox.Ok)
            return


    def restore_table_state(self, state):
        # 현재 테이블 상태를 초기화
        print(state)
        self.tbl_trans.clearContents()
        self.tbl_trans.setRowCount(state['row_count'])
        self.tbl_trans.setColumnCount(state['column_count'])
        for row, row_data in enumerate(state['data']):
            for col, cell_data in enumerate(row_data):
                item = QTableWidgetItem(cell_data)
                self.tbl_trans.setItem(row, col, item)

    def add_row(self):
        nowRowCount = self.tbl_trans.rowCount()
        self.tbl_trans.setRowCount(nowRowCount + 1)

    #def copy_row(self):

    def del_row(self):
        # 선택된 항목이 존재할 때 선택된 항목들에서 중복되지 않도록 Row를 추출(
        if len(self.selectedItems):
            for selectedItem in self.selectedItems:
                self.deleteRows.add(selectedItem.row())
        elif len(self.deleteRows):
            pass
        else:
            QMessageBox.warning(self, '선택된 행이 없습니다.', '행을 선택해주세요.', QMessageBox.Ok)
            return

        # 추출된 Row를 리스트로 바꿔서 역순으로 정렬(삭제중에 순서가 바뀌지 않게 하기 위해서)
        sorted_list = sorted(list(self.deleteRows), reverse=True)

        for row in sorted_list :
            self.tbl_trans.removeRow(row)

        self.selectedItems.clear()
        self.deleteRows.clear()

    def add_clicked_row(self):
        self.selectedItems = self.tbl_trans.selectedItems()
        print(self.selectedItems)

    def load_file_dialog(self):
        # 파일 다이얼로그 열기
        self.addEditAddFilePath, _ = self.file_dialog.getOpenFileName(self, "주소 지정 파일 가져오기", "",
                                                                      "All Files (*);;CSV Files (*.csv);;Excel Files (*.xlsx)")

        # 파일 경로를 QLabel에 표시
        if self.addEditAddFilePath:
            self.lbl_selected_file.setText(self.addEditAddFilePath.split('/')[-1])
        else:
            return

        # 파일을 데이터프레임으로 읽어오기
        df = pd.read_excel(self.addEditAddFilePath, skiprows=1,dtype=dict_dtype)
        df_drop_first_column = df.drop(df.columns[0], axis=1)
        df_drop_na = df_drop_first_column.dropna(axis=0)

        # 테이블 초기화
        self.tbl_trans.clear()
        self.tbl_trans.setRowCount(0)
        self.tbl_trans.setColumnCount(len(df.columns))
        # 새로운 열 개수
        new_column_count = len(add_cols)

        # 테이블 열 개수 변경
        self.tbl_trans.setColumnCount(new_column_count)

        # 새로운 열 라벨 설정
        self.tbl_trans.setHorizontalHeaderLabels(add_cols)

        # 테이블에 데이터 추가
        for row_index, row_data in df_drop_na.iterrows():
            # 새로운 행 추가
            row_position = self.tbl_trans.rowCount()
            self.tbl_trans.insertRow(row_position)

            # 각 셀에 데이터 추가
            for col_index, cell_data in enumerate(row_data):
                item = QTableWidgetItem(str(cell_data))
                self.tbl_trans.setItem(row_position, col_index, item)
        self.table_states.clear()
        self.save_table_state()

    def drop_load_file_event(self, event: QDropEvent):
        urls = [url.toLocalFile() for url in event.mimeData().urls()]
        #개수 검사
        if len(urls) > 1 :
            QMessageBox.warning(self, '너무 많은 개수의 파일 입력', '한가지 파일만 입력해주세요', QMessageBox.Ok)
            return

        if urls:
            url = urls[0]
            print("Dropped files:", url)
            #xlsx 파일인지 확인
            if not (url.split('.')[-1] == 'xlsx' or url.split('.')[-1] == 'csv'):
                QMessageBox.warning(self, '지정되지 않은 파일 입력', 'xlsx 파일이나 csv파일만 입력해주세요.', QMessageBox.Ok)
                return
            else:
                self.addEditAddFilePath = url
                self.lbl_selected_file.setText(self.addEditAddFilePath.split('/')[-1])

                # 파일을 데이터프레임으로 읽어오기
                df = pd.read_excel(self.addEditAddFilePath, skiprows=1, dtype=dict_dtype)
                df_drop_first_column = df.drop(df.columns[0], axis=1)
                df_drop_na = df_drop_first_column.dropna(axis=0)

                # 테이블 초기화
                self.tbl_trans.clear()
                self.tbl_trans.setRowCount(0)
                self.tbl_trans.setColumnCount(len(df.columns))
                # 새로운 열 개수
                new_column_count = len(add_cols)

                # 테이블 열 개수 변경
                self.tbl_trans.setColumnCount(new_column_count)

                # 새로운 열 라벨 설정
                self.tbl_trans.setHorizontalHeaderLabels(add_cols)

                # 테이블에 데이터 추가
                for row_index, row_data in df_drop_na.iterrows():
                    # 새로운 행 추가
                    row_position = self.tbl_trans.rowCount()
                    self.tbl_trans.insertRow(row_position)

                    # 각 셀에 데이터 추가
                    for col_index, cell_data in enumerate(row_data):
                        item = QTableWidgetItem(str(cell_data))
                        self.tbl_trans.setItem(row_position, col_index, item)
                self.table_states.clear()
                self.save_table_state()

    def save_to_excel(self):
        file_path, _ = self.file_dialog.getSaveFileName(self, "Excel 파일로 저장", "", "Excel Files (*.xlsx)")

        if file_path:
            # 행 인덱스를 숫자로 설정하여 데이터프레임 생성
            wb = load_workbook("./address_header.xlsx")

            # 기존 시트 선택 또는 새 시트 생성
            ws = wb.active
            for row in range(self.tbl_trans.rowCount()):
                row_data = [row + 1]  # 인덱스는 1부터 시작
                row_data.extend([self.tbl_trans.item(row, col).text() for col in range(self.tbl_trans.columnCount())])
                ws.append(row_data)

            wb.save(file_path)


    def ex_output_name(self):
        realName = self.transOutFilePath.split('/')[-1].split('.')[0]
        directory_path = os.path.dirname(self.transOutFilePath)
        files = os.listdir(directory_path)
        files = [file_name for file_name in files if '.xlsx' in file_name]
        files = [file_name for file_name in files if realName in file_name]
        files.sort()

        self.num = 0
        for file in files:
            file = file.replace(realName, "").split(".")[0]
            text = self.create_suffix().split(".")[0]
            if file == text:
                self.num += 1

        self.lbl_ex_output_name.setText('생성될 예정의 출력 파일명은 ' + realName + self.create_suffix() + ' 입니다.')

    def change_suffix(self):
        suffix = self.txt_suffix.toPlainText()
        if suffix != '' :
            #특수문자 제거
            self.suffix = re.sub('[\/:*?"<>|]','',suffix)
            self.set_suffix_lbl()
        else:
            QMessageBox.warning(self, '입력된 값이 없습니다.', '값을 입력해주세요', QMessageBox.Ok)

    def change_sufNum(self):
        self.sufNum = 0
        self.set_suffix_lbl()

    def change_sufNum_2(self):
        self.sufNum = 1
        self.set_suffix_lbl()

    def change_sufNum_3(self):
        self.sufNum = 2
        self.set_suffix_lbl()

    def create_suffix(self):
        suffix = '_'+self.suffix + (self.suffixNum[self.sufNum] % self.num) + '.xlsx'
        return suffix

    def set_suffix_lbl(self):
        self.lbl_save_name_txt.setText(self.create_suffix() + ' 입니다.')
        if self.transOutFilePath != '':
            self.ex_output_name()


def resource_path(*relative_Path_AND_File):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = getattr(
            sys,
            '_MEIPASS',
            os.path.dirname(os.path.abspath(__file__))
        )
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, *relative_Path_AND_File)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    sys.exit(app.exec_())
